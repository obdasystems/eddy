# -*- coding: utf-8 -*-
##########################################################################
#                                                                        #
#  Eddy: a graphical editor for the specification of Graphol ontologies  #
#  Copyright (C) 2015 Daniele Pantaleone <danielepantaleone@me.com>      #
#                                                                        #
#  This program is free software: you can redistribute it and/or modify  #
#  it under the terms of the GNU General Public License as published by  #
#  the Free Software Foundation, either version 3 of the License, or     #
#  (at your option) any later version.                                   #
#                                                                        #
#  This program is distributed in the hope that it will be useful,       #
#  but WITHOUT ANY WARRANTY; without even the implied warranty of        #
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the          #
#  GNU General Public License for more details.                          #
#                                                                        #
#  You should have received a copy of the GNU General Public License     #
#  along with this program. If not, see <http://www.gnu.org/licenses/>.  #
#                                                                        #
#  #####################                          #####################  #
#                                                                        #
#  Graphol is developed by members of the DASI-lab group of the          #
#  Dipartimento di Ingegneria Informatica, Automatica e Gestionale       #
#  A.Ruberti at Sapienza University of Rome: http://www.dis.uniroma1.it  #
#                                                                        #
#     - Domenico Lembo <lembo@dis.uniroma1.it>                           #
#     - Valerio Santarelli <santarelli@dis.uniroma1.it>                  #
#     - Domenico Fabio Savo <savo@dis.uniroma1.it>                       #
#     - Daniele Pantaleone <pantaleone@dis.uniroma1.it>                  #
#     - Marco Console <console@dis.uniroma1.it>                          #
#                                                                        #
##########################################################################

from abc import abstractmethod
import csv

from PyQt5 import (
    QtCore,
    QtGui,
    QtWidgets,
)
import openpyxl

from eddy.core.commands.iri import (
    CommandIRIAddAnnotationAssertion,
    CommandIRIRemoveAnnotationAssertion,
)
from eddy.core.commands.project import CommandProjectAddAnnotationProperty
from eddy.core.datatypes.graphol import Item
from eddy.core.datatypes.system import File
from eddy.core.loaders.common import AbstractOntologyLoader
from eddy.core.output import getLogger
from eddy.core.owl import (
    AnnotationAssertion,
    IllegalNamespaceError,
)
from eddy.ui.progress import BusyProgressDialog

LOGGER = getLogger()


class TemplateLoader(AbstractOntologyLoader):
    """
    Extends AbstractOntologyLoader with facilities to load annotations from CSV/Xlsx file format.
    """

    def __init__(self, path, project, session):
        """
        Initialize the Template importer.
        :type path: str
        :type project: Project
        :type session: Session
        """
        super().__init__(path, project, session)

    @classmethod
    @abstractmethod
    def filetype(cls):
        """
        Returns the type of the file that will be used for the import.
        :return: File
        """
        pass

    @abstractmethod
    def run(self):
        """
        Perform the load of the ontology and the merge with the current project.
        """
        pass

    def importAnnotations(self, rows, override):
        """
        Import the annotations.
        """
        types = {
            'Data Property': Item.AttributeNode,
            'Class': Item.ConceptNode,
            'Named Individual': Item.IndividualNode,
            'Object Property': Item.RoleNode
        }
        for row in rows:
            resource = str(row[0]) if len(row) > 0 else None
            if resource and resource != 'None':
                QtCore.QCoreApplication.processEvents()

                simple_name = str(row[1])
                type = str(row[2])
                annotation = str(row[3])
                datatype = str(row[4])
                lang = str(row[5]) if str(row[5]) != '' else 'en'
                value = str(row[6])

                resourceIRI = self.project.getIRI(resource)
                annotationIRI = self.project.getIRI(annotation)
                datatypeIRI = None
                if datatype:
                    datatypeIRI = self.project.getIRI(datatype)

                if value == '' or value == 'None':
                    pass
                else:
                    # ADD ANNOTATION PROPERTY
                    try:
                        QtCore.QCoreApplication.processEvents()

                        if annotationIRI not in self.project.getAnnotationPropertyIRIs():
                            self.project.isValidIdentifier(annotation)

                            command = CommandProjectAddAnnotationProperty(self.project, annotation)
                            self.session.undostack.beginMacro(
                                'Add annotation property {0} '.format(annotation))
                            if command:
                                self.session.undostack.push(command)
                            self.session.undostack.endMacro()

                    except IllegalNamespaceError as e:
                        # noinspection PyArgumentList
                        msgBox = QtWidgets.QMessageBox(
                            QtWidgets.QMessageBox.Warning,
                            'Entity Definition Error',
                            'Illegal namespace defined.',
                            informativeText='The string "{}" is not a legal IRI'.format(annotation),
                            detailedText=str(e),
                            parent=self,
                        )
                        msgBox.exec_()

                    # CREATE ANNOTATION ASSERTION
                    annotationAss = AnnotationAssertion(resourceIRI, annotationIRI, value,
                                                        type=datatypeIRI,
                                                        language=lang)

                    if override:
                        QtCore.QCoreApplication.processEvents()

                        assertionSub = annotationAss.subject
                        assertionProp = annotationAss.assertionProperty
                        assertionLang = annotationAss.language
                        processed = set()
                        for item in self.project.iriOccurrences():
                            QtCore.QCoreApplication.processEvents()

                            # LOOK FOR RESOURCE
                            if item.iri not in processed:
                                if item.isNode() and item.type() == types[
                                    type] and item.iri is assertionSub:
                                    existing = []
                                    # LOOK FOR ANNOTATION PROPERTY
                                    for prop, assertsList in item.iri.annotationAssertionMapItems:
                                        if prop is assertionProp:
                                            # LOOK FOR LANGUAGE
                                            currList = assertsList
                                            for assertion in currList:
                                                if assertion.language == assertionLang:
                                                    existing.append(assertion)

                                    # REMOVE  ALL EXISTING ANNOTATION ASSERTIONS
                                    for annAssertion in existing:
                                        QtCore.QCoreApplication.processEvents()

                                        self.session.undostack.push(
                                            CommandIRIRemoveAnnotationAssertion(
                                                self.project, item.iri,
                                                annAssertion))

                                    # ADD NEW ANNOTATION ASSERTION FOR IRI-PROPERTY-LANG
                                    self.session.undostack.push(
                                        CommandIRIAddAnnotationAssertion(self.project,
                                                                         item.iri, annotationAss))
                                    processed.add(item.iri)

                    else:

                        processed = set()
                        for item in self.project.iriOccurrences():
                            QtCore.QCoreApplication.processEvents()

                            if item.iri not in processed:

                                if item.isNode() and item.type() == types[
                                    type] and item.iri is annotationAss.subject:
                                    # ADD ANNOTATION ASSERTION
                                    self.session.undostack.push(
                                        CommandIRIAddAnnotationAssertion(self.project, item.iri,
                                                                         annotationAss))
                                    processed.add(item.iri)


class CsvLoader(TemplateLoader):
    """
    Extends AbstractOntologyLoader with facilities to load annotations from CSV file format.
    """

    def __init__(self, path, project, session):
        """
        Initialize the Csv importer.
        :type path: str
        :type project: Project
        :type session: Session
        """
        super().__init__(path, project, session)

    @classmethod
    def filetype(cls):
        """
        Returns the type of the file that will be used for the import.
        :return: File
        """
        return File.Csv

    def run(self, file, override):
        """
        Gets the set of annotations from CSV file.
        """
        bp = BusyProgressDialog('Loading Annotations')
        bp.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        with bp:
            file = open(file)
            csvreader = csv.reader(file)
            QtCore.QCoreApplication.processEvents()

            header = next(csvreader)
            checkArray = ['RESOURCE', 'SIMPLE_NAME', 'TYPE', 'ANNOTATION', 'DATATYPE', 'LANG',
                          'VALUE']
            for i in range(0, len(header)):
                if checkArray[i] in header[i]:
                    continue
                else:
                    msgbox = QtWidgets.QMessageBox()
                    msgbox.setIconPixmap(QtGui.QIcon(':/icons/48/ic_warning_black').pixmap(48))
                    msgbox.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
                    msgbox.setWindowTitle('Template Mismatch')
                    msgbox.setText('The imported file does not match the predifined template.\n'
                                   'Please fill the predifined template and try again.')
                    msgbox.setTextFormat(QtCore.Qt.RichText)
                    msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msgbox.exec_()
                    return

            rows = []
            for row in csvreader:
                QtCore.QCoreApplication.processEvents()
                rows.append(row)
            file.close()
            self.importAnnotations(rows, override)


class XlsxLoader(TemplateLoader):
    """
    Extends AbstractOntologyLoader with facilities to load annotations from Xlsx file format.
    """

    def __init__(self, path, project, session):
        """
        Initialize the Xlsx importer.
        :type path: str
        :type project: Project
        :type session: Session
        """
        super().__init__(path, project, session)

    @classmethod
    def filetype(cls):
        """
        Returns the type of the file that will be used for the import.
        :return: File
        """
        return File.Xlsx

    def run(self, file, override):
        """
        Gets the set of annotations from Xlsx file.
        """
        bp = BusyProgressDialog('Loading Annotations')
        bp.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
        with bp:
            workbook = openpyxl.load_workbook(file, enumerate)
            sheet = workbook.worksheets[0]
            QtCore.QCoreApplication.processEvents()

            number_of_rows = sheet.max_row + 1
            number_of_columns = sheet.max_column + 1

            rows = []
            breakFlag = False
            for i in range(1, number_of_rows):
                row = []
                for x in range(1, 8):
                    QtCore.QCoreApplication.processEvents()
                    value = sheet.cell(row=i, column=x).value
                    if x == 1 and (value == 'None' or value is None):
                        breakFlag = True
                        break
                    row.append(value)
                if breakFlag:
                    break
                rows.append(row)

            header = rows[0]
            checkArray = ['RESOURCE', 'SIMPLE_NAME', 'TYPE', 'ANNOTATION','DATATYPE', 'LANG', 'VALUE']
            for i in range(0, len(header)):
                if checkArray[i] in header[i]:
                    continue
                else:
                    msgbox = QtWidgets.QMessageBox()
                    msgbox.setIconPixmap(QtGui.QIcon(':/icons/48/ic_warning_black').pixmap(48))
                    msgbox.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
                    msgbox.setWindowTitle('Template Mismatch')
                    msgbox.setText('The imported file does not match the predifined template.\n'
                                   'Please fill the predifined template and try again.')
                    msgbox.setTextFormat(QtCore.Qt.RichText)
                    msgbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                    msgbox.exec_()
                    return

            rows = rows[1:]
            self.importAnnotations(rows, override)
