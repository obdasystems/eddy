# -*- coding: utf-8 -*-

##########################################################################
#                                                                        #
#  Eddy: a graphical editor for the specification of Graphol ontologies  #
#  Copyright (C) 2015 Daniele Pantaleone <danielepantaleone@me.com>      #
#                                                                        #
#  This program is free software: you can redistribute it and/or modify  #
#  it under the terms of the GNU General Public License as published by  #
#  the Free Software Foundation, either version 3 of the License, or     #
#  (at your option) any later version.                                   #
#                                                                        #
#  This program is distributed in the hope that it will be useful,       #
#  but WITHOUT ANY WARRANTY; without even the implied warranty of        #
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the          #
#  GNU General Public License for more details.                          #
#                                                                        #
#  You should have received a copy of the GNU General Public License     #
#  along with this program. If not, see <http://www.gnu.org/licenses/>.  #
#                                                                        #
#  #####################                          #####################  #
#                                                                        #
#  Graphol is developed by members of the DASI-lab group of the          #
#  Dipartimento di Ingegneria Informatica, Automatica e Gestionale       #
#  A.Ruberti at Sapienza University of Rome: http://www.dis.uniroma1.it  #
#                                                                        #
#     - Domenico Lembo <lembo@dis.uniroma1.it>                           #
#     - Valerio Santarelli <santarelli@dis.uniroma1.it>                  #
#     - Domenico Fabio Savo <savo@dis.uniroma1.it>                       #
#     - Daniele Pantaleone <pantaleone@dis.uniroma1.it>                  #
#     - Marco Console <console@dis.uniroma1.it>                          #
#                                                                        #
##########################################################################


from __future__ import annotations

from abc import ABCMeta, abstractmethod
import csv
import io
from copy import copy
from typing import (
    cast,
    Any,
    Dict,
    List,
    Set,
    TYPE_CHECKING,
)

from PyQt5 import (
    QtCore,
    QtGui,
    QtWidgets,
)

from openpyxl import Workbook
from openpyxl.styles import DEFAULT_FONT, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

from eddy.core.common import HasWidgetSystem
from eddy.core.datatypes.graphol import Item
from eddy.core.datatypes.system import File
from eddy.core.exporters.common import AbstractProjectExporter
from eddy.core.functions.fsystem import fwrite
from eddy.core.functions.path import openPath
from eddy.core.functions.signals import connect
from eddy.core.output import getLogger
from eddy.core.owl import Annotation, AnnotationAssertionProperty
from eddy.ui.dialogs import DiagramSelectionDialog
from eddy.ui.fields import (
    CheckBox,
    StringField,
)

if TYPE_CHECKING:
    from eddy.core.project import Project
    from eddy.ui.session import Session

LOGGER = getLogger()


class AbstractMetadataExporter(AbstractProjectExporter):
    """
    Abstract base class for exporting Graphol projects' metadata.
    """
    __metaclass__ = ABCMeta

    KeyResource = 'RESOURCE'
    KeySimpleName = 'SIMPLE_NAME'
    KeyType = 'TYPE'
    KeyAnnotation = 'ANNOTATION'
    KeyDataType = 'DATATYPE'
    KeyLang = 'LANG'
    KeyValue = 'VALUE'
    Types = {
        Item.AttributeNode: 'Data Property',
        Item.ConceptNode: 'Class',
        Item.IndividualNode: 'Named Individual',
        Item.RoleNode: 'Object Property',
    }

    def __init__(self, project: Project, session: Session = None, **kwargs: Any) -> None:
        """
        Initialize the metadata exporter.
        """
        super().__init__(project, session)
        self.diagrams = kwargs.get('diagrams', None)
        self.annotations = kwargs.get('annotations', None)
        self.items = kwargs.get('items', None)
        self.open = kwargs.get('open', False)
        self.includeEntitiesWithoutAnnotations = kwargs.get('includeEntitiesWithoutAnnotations',False)


    #############################################
    #   INTERFACE
    #################################

    def metadataHeader(self) -> List[str]:
        """
        Returns a list containing the metadata table column names.
        """
        return [
            self.KeyResource,
            self.KeySimpleName,
            self.KeyType,
            self.KeyAnnotation,
            self.KeyDataType,
            self.KeyLang,
            self.KeyValue,
        ]

    def metadata(self) -> List[Dict[str, str]]:
        """
        Returns a list containing the metadata table rows, as dictionaries
        indexed by their corresponding column name.
        """
        if self.diagrams is None:
            self.diagrams = self.project.diagrams()
        if self.annotations is None:
            self.annotations = self.project.getAnnotationPropertyIRIs()
        if self.items is None:
            self.items = self.Types.keys()
        meta = []  # type: List[Dict[str, str]]
        processed = set()
        # PROJECT METADATA
        for diagram in self.diagrams:
            for node in self.project.iriOccurrences(diagram=diagram):
                if node.type() not in self.items or node.iri in processed:
                    continue
                if self.includeEntitiesWithoutAnnotations and len(
                    node.iri.annotationAssertions) == 0:
                    meta.append({
                        self.KeyResource: str(node.iri),
                        self.KeySimpleName: node.iri.getSimpleName(),
                        self.KeyType: self.Types.get(node.type()),
                        self.KeyAnnotation: '',
                        self.KeyDataType: '',
                        self.KeyLang: '',
                        self.KeyValue: '',
                    })
                for annotation in node.iri.annotationAssertions:
                    if annotation.assertionProperty in self.annotations:
                        meta.append({
                            self.KeyResource: str(node.iri),
                            self.KeySimpleName: node.iri.getSimpleName(),
                            self.KeyType: self.Types.get(node.type()),
                            self.KeyAnnotation: annotation.assertionProperty,
                            self.KeyDataType: annotation.datatype or '',
                            self.KeyLang: annotation.language or '',
                            self.KeyValue: str(annotation.value),
                        })
                processed.add(node.iri)

        # IMPORTED METADATA
        for ont in self.project.importedOntologies:
            resources = []
            if Item.ConceptNode in self.items:
                resources.extend(map(lambda v: ('Class', v), ont.classes))
            if Item.RoleNode in self.items:
                resources.extend(map(lambda v: ('Object Property', v), ont.objectProperties))
            if Item.AttributeNode in self.items:
                resources.extend(map(lambda v: ('Data Property', v), ont.dataProperties))
            if Item.IndividualNode in self.items:
                resources.extend(map(lambda v: ('Named Individual', v), ont.individuals))
            for entityType, resource in resources:
                if resource in processed:
                    continue
                for annotation in resource.annotationAssertions:
                    if annotation.assertionProperty in self.annotations:
                        meta.append({
                            self.KeyResource: str(resource),
                            self.KeySimpleName: resource.getSimpleName(),
                            self.KeyType: entityType,
                            self.KeyAnnotation: annotation.assertionProperty,
                            self.KeyDataType: annotation.datatype or '',
                            self.KeyLang: annotation.language or '',
                            self.KeyValue: str(annotation.value),
                        })
                processed.add(resource)

        return sorted(meta, key=lambda i: i[self.KeyResource])

    @classmethod
    @abstractmethod
    def filetype(cls) -> File:
        """
        Returns the type of the file that will be used for the export.
        """
        pass

    @abstractmethod
    def run(self, path: str) -> None:
        """
        Perform the export.
        """
        pass

class CsvProjectExporter(AbstractMetadataExporter):
    """
    This class can be used to export Graphol projects into CSV format.
    """
    #############################################
    #   INTERFACE
    #################################

    @classmethod
    def filetype(cls) -> File:
        """
        Returns the type of the file that will be used for the export.
        """
        return File.Csv

    def run(self, path: str) -> None:
        """
        Perform CSV file generation.
        """
        LOGGER.info('Exporting project metadata %s to CSV file: %s', self.project.name, path)

        # SELECT PROJECT DIAGRAMS
        if self.diagrams is None:
            if not self.project.isEmpty():
                dialog = DiagramSelectionDialog(self.session)
                if not dialog.exec_():
                    return
                self.diagrams = dialog.selectedDiagrams()
            else:
                self.diagrams = self.project.diagrams()

        # SELECT PROJECT ANNOTATIONS
        if self.annotations is None or self.items is None:
            dialog = AnnotationSelectionDialog(self.project, parent=self.session)
            if not dialog.exec_():
                return
            self.annotations = dialog.selectedAnnotations()
            self.items = dialog.selectedItems()
            # CHECK INCLUSION OF ENTITIES WITHOUT ANNOTATIONS
            self.includeEntitiesWithoutAnnotations = dialog.checked()

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        writer.writerow(self.metadataHeader())
        writer.writerows(map(lambda row: row.values(), self.metadata()))
        fwrite(buffer.getvalue(), path, newline='')
        if self.open:
            openPath(path)

class XlsxProjectExporter(AbstractMetadataExporter):
    """
    This class can be used to export Graphol projects into Excel 2007+ .xlsx format.
    """
    #############################################
    #   INTERFACE
    #################################

    @classmethod
    def filetype(cls) -> File:
        """
        Returns the type of the file that will be used for the export.
        """
        return File.Xlsx

    def run(self, path: str) -> None:
        """
        Perform Xlsx file generation.
        """
        LOGGER.info('Exporting project metadata %s to XLSX file: %s', self.project.name, path)

        # SELECT PROJECT DIAGRAMS
        if self.diagrams is None:
            if not self.project.isEmpty():
                dialog = DiagramSelectionDialog(self.session)
                if not dialog.exec_():
                    return
                self.diagrams = dialog.selectedDiagrams()
            else:
                self.diagrams = self.project.diagrams()

        # SELECT PROJECT ANNOTATIONS
        if self.annotations is None or self.items is None:
            dialog = AnnotationSelectionDialog(self.project, parent=self.session)
            if not dialog.exec_():
                return
            self.annotations = dialog.selectedAnnotations()
            self.items = dialog.selectedItems()
            # CHECK INCLUSION OF ENTITIES WITHOUT ANNOTATIONS
            self.includeEntitiesWithoutAnnotations = dialog.checked()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.project.name
        bodyFont = copy(DEFAULT_FONT)
        headFont = copy(bodyFont)
        headFont.bold = True
        alignment = Alignment(vertical='center', wrapText=True)
        headStyle = NamedStyle(name='head', font=headFont, alignment=alignment)
        bodyStyle = NamedStyle(name='body', font=bodyFont, alignment=alignment)
        # HEADER ROW
        for j, title in enumerate(self.metadataHeader(), start=1):
            worksheet.cell(row=1, column=j, value=title).style = headStyle
        worksheet.freeze_panes = 'A2'
        # METADATA ROWS
        metadata = self.metadata()
        if metadata:
            for i, row in enumerate(metadata, start=2):
                for j, value in enumerate(row.values(), start=1):
                    worksheet.cell(row=i, column=j, value=str(value)).style = bodyStyle
        # AUTOFIT COLUMN WIDTHS
        for j, key in enumerate(self.metadataHeader(), start=1):
            # Length of the longest line in a multi-line string
            def maxlen(s): return max(map(len, str(s).split('\n')))

            # Compute the column width as the max between the length of
            # the column header, and the max length of all column values.
            width = max(len(key), max(map(lambda d: maxlen(d[key]), metadata)))
            worksheet.column_dimensions[get_column_letter(j)].width = width
        # AUTOFIT ROW HEIGHTS
        for i, row in enumerate(metadata, start=2):
            height = max(map(lambda s: len(s.split('\n')), map(str, row.values())))
            # Set height as number of lines x default font height + some padding
            worksheet.row_dimensions[i].height = height * bodyFont.size * 1.25
        workbook.save(path)
        if self.open:
            openPath(path)


class AbstractTemplateExporter(AbstractProjectExporter):

    __metaclass__ = ABCMeta

    KeyResource = 'RESOURCE'
    KeySimpleName = 'SIMPLE_NAME'
    KeyType = 'TYPE'
    KeyAnnotation = 'ANNOTATION'
    KeyDataType = 'DATATYPE'
    KeyLang = 'LANG'
    KeyValue = 'VALUE'
    Types = {
        Item.AttributeNode: 'Data Property',
        Item.ConceptNode: 'Class',
        Item.IndividualNode: 'Named Individual',
        Item.RoleNode: 'Object Property',
    }

    def __init__(self, project, session, **kwargs: Any):
        """
        Initialize the Template exporter.
        :type project: Project
        :type session: Session
        """
        super().__init__(project, session)
        self.diagrams = kwargs.get('diagrams', None)
        self.items = kwargs.get('items', None)
        self.open = kwargs.get('open', False)


    @classmethod
    @abstractmethod
    def filetype(cls):
        """
        Returns the type of the file that will be used for the import.
        :return: File
        """
        pass

    def dataHeader(self) -> List[str]:
        """
        Returns a list containing the metadata table column names.
        """
        return [
            self.KeyResource,
            self.KeySimpleName,
            self.KeyType,
            self.KeyAnnotation,
            self.KeyDataType,
            self.KeyLang,
            self.KeyValue,
        ]

    def openDialogs(self):

        if self.diagrams is None:
            if not self.project.isEmpty():
                dialog = DiagramSelectionDialog(self.session)
                if not dialog.exec_():
                    return
                self.diagrams = dialog.selectedDiagrams()
            else:
                self.diagrams = self.project.diagrams()

        if self.items is None:
            dialog = EntityTypesSelectionDialog(self.project, parent=self.session)
            if not dialog.exec_():
                return
            self.items = dialog.selectedItems()

    def data(self):
        meta = []
        processed = []
        for diagram in self.diagrams:
            for item in diagram.items():
                if item.type() in self.items and str(item.iri) not in processed:
                    meta.append({
                        self.KeyResource: str(item.iri),
                        self.KeySimpleName: item.iri.getSimpleName(),
                        self.KeyType: self.Types.get(item.type()),
                        self.KeyAnnotation: AnnotationAssertionProperty.Label.value,
                        self.KeyDataType: '',
                        self.KeyLang: '',
                        self.KeyValue: '',
                    })
                    processed.append(str(item.iri))

        return sorted(meta, key=lambda i: i[self.KeyResource])

    @abstractmethod
    def run(self, path: str) -> None:
        """
        Perform the export.
        """
        pass

class CsvTemplateExporter(AbstractTemplateExporter):

    @classmethod
    def filetype(cls) -> File:
        """
        Returns the type of the file that will be used for the export.
        """
        return File.Csv

    def run(self, path: str) -> None:
        """
        Perform CSV file generation.
        """
        LOGGER.info('Exporting metadata template %s to CSV file: %s', self.project.name, path)

        self.openDialogs()

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        writer.writerow(self.dataHeader())

        meta = self.data()
        writer.writerows(map(lambda row: row.values(), meta))
        fwrite(buffer.getvalue(), path, newline='')
        if self.open:
            openPath(path)

class XlsxTemplateExporter(AbstractTemplateExporter):
    """
    This class can be used to export Graphol projects into Excel 2007+ .xlsx format.
    """
    #############################################
    #   INTERFACE
    #################################

    @classmethod
    def filetype(cls) -> File:
        """
        Returns the type of the file that will be used for the export.
        """
        return File.Xlsx

    def run(self, path: str) -> None:
        """
        Perform Xlsx file generation.
        """
        LOGGER.info('Exporting metadata template %s to XLSX file: %s', self.project.name, path)

        self.openDialogs()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.project.name
        bodyFont = copy(DEFAULT_FONT)
        headFont = copy(bodyFont)
        headFont.bold = True
        alignment = Alignment(vertical='center', wrapText=True)
        headStyle = NamedStyle(name='head', font=headFont, alignment=alignment)
        bodyStyle = NamedStyle(name='body', font=bodyFont, alignment=alignment)
        # HEADER ROW
        for j, title in enumerate(self.dataHeader(), start=1):
            worksheet.cell(row=1, column=j, value=title).style = headStyle
        worksheet.freeze_panes = 'A2'
        # METADATA ROWS
        metadata = self.data()
        if metadata:
            for i, row in enumerate(metadata, start=2):
                for j, value in enumerate(row.values(), start=1):
                    worksheet.cell(row=i, column=j, value=str(value)).style = bodyStyle
        # AUTOFIT COLUMN WIDTHS
        for j, key in enumerate(self.dataHeader(), start=1):
            # Length of the longest line in a multi-line string
            def maxlen(s): return max(map(len, str(s).split('\n')))

            # Compute the column width as the max between the length of
            # the column header, and the max length of all column values.
            width = max(len(key), max(map(lambda d: maxlen(d[key]), metadata)))
            worksheet.column_dimensions[get_column_letter(j)].width = width
        # AUTOFIT ROW HEIGHTS
        for i, row in enumerate(metadata, start=2):
            height = max(map(lambda s: len(s.split('\n')), map(str, row.values())))
            # Set height as number of lines x default font height + some padding
            worksheet.row_dimensions[i].height = height * bodyFont.size * 1.25
        workbook.save(path)
        if self.open:
            openPath(path)


class EntityTypesSelectionDialog(HasWidgetSystem, QtWidgets.QDialog):
    """
    Extends `QtWidgets.QDialog` providing a form for selecting a subset
    of the project annotation properties.
    """

    def __init__(
        self,
        project: 'Project',
        parent: QtWidgets.QWidget = None,
        **kwargs: Any
    ) -> None:
        """
        Initialize the dialog.
        """
        super().__init__(parent=parent, **kwargs)
        self._project = project
        self.diagrams = kwargs.get('diagrams', None)
        self.items = kwargs.get('items', {
            Item.ConceptNode,
            Item.RoleNode,
            Item.AttributeNode,
            Item.IndividualNode,
        })


        # ENTITY TYPES GROUP BOX
        groupbox = QtWidgets.QGroupBox('Entity Types')
        groupbox.setObjectName('entities_group')
        layout = QtWidgets.QHBoxLayout(groupbox)
        for item in self.items:
            checkbox = CheckBox(AbstractMetadataExporter.Types.get(item), self)
            checkbox.setObjectName(item.shortName)
            checkbox.setChecked(True)
            layout.addWidget(checkbox)
            self.addWidget(checkbox)
        self.addWidget(groupbox)


        # CONFIRMATION BOX
        confirmation = QtWidgets.QDialogButtonBox(QtCore.Qt.Horizontal, self)
        confirmation.setObjectName('confirmation_box')
        confirmation.addButton(QtWidgets.QDialogButtonBox.Cancel)
        confirmation.addButton(QtWidgets.QDialogButtonBox.Ok)
        connect(confirmation.accepted, self.accept)
        connect(confirmation.rejected, self.reject)
        self.addWidget(confirmation)

        # DIALOG LAYOUT
        mainLayout = QtWidgets.QVBoxLayout(self)
        mainLayout.setAlignment(QtCore.Qt.AlignTop)
        mainLayout.addWidget(self.widget('entities_group'))
        mainLayout.addWidget(self.widget('confirmation_box'), 0, QtCore.Qt.AlignRight)

        self.setMinimumSize(640, 100)
        self.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
        self.setWindowTitle("Select Entity Types")

    #############################################
    #   PROPERTIES
    #################################

    @property
    def project(self) -> 'Project':
        """
        Returns the active project.
        """
        return self._project

    #############################################
    #   INTERFACE
    #################################


    def selectedItems(self) -> Set[Item]:
        """
        Return the set of selected item types.
        """
        items = set()
        for item in self.items:
            checkbox = cast(QtWidgets.QCheckBox, self.widget(item.shortName))
            if checkbox.isChecked():
                items.add(item)
        return items

class AnnotationSelectionDialog(HasWidgetSystem, QtWidgets.QDialog):
    """
    Extends `QtWidgets.QDialog` providing a form for selecting a subset
    of the project annotation properties.
    """

    def __init__(
        self,
        project: Project,
        parent: QtWidgets.QWidget = None,
        **kwargs: Any
    ) -> None:
        """
        Initialize the dialog.
        """
        super().__init__(parent=parent, **kwargs)
        self._project = project
        self.diagrams = kwargs.get('diagrams', None)
        self.items = kwargs.get('items', {
            Item.ConceptNode,
            Item.RoleNode,
            Item.AttributeNode,
            Item.IndividualNode,
        })

        # ANNOTATION LIST WIDGET
        listview = AnnotationListWidget(self.project, self)
        listview.setObjectName('annotations_list')
        self.addWidget(listview)

        # LIST WIDGET GROUP BOX
        groupbox = QtWidgets.QGroupBox('Annotation Properties')
        groupbox.setObjectName('annotations_group')
        layout = QtWidgets.QVBoxLayout(groupbox)
        layout.setAlignment(QtCore.Qt.AlignTop)
        layout.addWidget(self.widget('annotations_list'))
        self.addWidget(groupbox)

        # ENTITY TYPES GROUP BOX
        groupbox = QtWidgets.QGroupBox('Entity Types')
        groupbox.setObjectName('entities_group')
        layout = QtWidgets.QHBoxLayout(groupbox)
        for item in self.items:
            checkbox = CheckBox(AbstractMetadataExporter.Types.get(item), self)
            checkbox.setObjectName(item.shortName)
            checkbox.setChecked(True)
            layout.addWidget(checkbox)
            self.addWidget(checkbox)
        self.addWidget(groupbox)

        # ENTITY WITHOUT ANNOTATIONS CHECKBOX
        groupbox2 = QtWidgets.QGroupBox('Entities Without Annotations')
        groupbox2.setObjectName('entities_without_annotations')
        layout2 = QtWidgets.QHBoxLayout(groupbox2)

        checkbox = CheckBox('Include entities without annotations')
        checkbox.setObjectName('all_entities_checkbox')
        layout2.addWidget(checkbox)
        self.addWidget(checkbox)

        self.addWidget(groupbox2)

        # CONFIRMATION BOX
        confirmation = QtWidgets.QDialogButtonBox(QtCore.Qt.Horizontal, self)
        confirmation.setObjectName('confirmation_box')
        confirmation.addButton(QtWidgets.QDialogButtonBox.Cancel)
        confirmation.addButton(QtWidgets.QDialogButtonBox.Ok)
        connect(confirmation.accepted, self.accept)
        connect(confirmation.rejected, self.reject)
        self.addWidget(confirmation)

        # DIALOG LAYOUT
        mainLayout = QtWidgets.QVBoxLayout(self)
        mainLayout.setAlignment(QtCore.Qt.AlignTop)
        mainLayout.addWidget(self.widget('annotations_group'))
        mainLayout.addWidget(self.widget('entities_group'))
        mainLayout.addWidget(self.widget('entities_without_annotations'))
        mainLayout.addWidget(self.widget('confirmation_box'), 0, QtCore.Qt.AlignRight)

        self.setMinimumSize(640, 480)
        self.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
        self.setWindowTitle("Select Annotation Properties")

    #############################################
    #   PROPERTIES
    #################################

    @property
    def project(self) -> Project:
        """
        Returns the active project.
        """
        return self._project

    #############################################
    #   INTERFACE
    #################################

    def selectedAnnotations(self) -> Set[Annotation]:
        """
        Returns the set of selected annotations.
        """
        return cast(AnnotationListWidget, self.widget('annotations_list')).selectedAnnotations()

    def selectedItems(self) -> Set[Item]:
        """
        Return the set of selected item types.
        """
        items = set()
        for item in self.items:
            checkbox = cast(QtWidgets.QCheckBox, self.widget(item.shortName))
            if checkbox.isChecked():
                items.add(item)
        return items

    def checked(self):

        checked = self.widget('all_entities_checkbox').isChecked()
        return checked


class AnnotationListWidget(HasWidgetSystem, QtWidgets.QWidget):
    """
    This class implements a widget that can be used to select a subset
    of the project's annotation properties.
    """
    AnnotationRole = QtCore.Qt.UserRole + 101

    def __init__(
        self,
        project: Project,
        parent: QtWidgets.QWidget = None,
        **kwargs: Any
    ) -> None:
        """
        Initialize the widget.
        """
        super().__init__(parent=parent, **kwargs)
        self._project = project

        self.model = QtGui.QStandardItemModel(self)
        self.proxy = QtCore.QSortFilterProxyModel(self)
        self.proxy.setDynamicSortFilter(False)
        self.proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.proxy.setSortCaseSensitivity(QtCore.Qt.CaseSensitive)
        self.proxy.setSourceModel(self.model)

        # SEARCH FIELD
        search = StringField(self, objectName='search_field')
        search.setPlaceholderText('Search...')
        search.setClearButtonEnabled(True)
        connect(search.textChanged, self.doFilterItem)
        connect(search.returnPressed, self.onReturnPressed)
        self.addWidget(search)

        # LIST WIDGET BOX
        listview = QtWidgets.QListView(self)
        listview.setObjectName('annotations_list')
        listview.setModel(self.proxy)
        listview.setAlternatingRowColors(True)
        listview.setContextMenuPolicy(QtCore.Qt.PreventContextMenu)
        listview.setDragDropMode(QtWidgets.QAbstractItemView.NoDragDrop)
        listview.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        listview.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        listview.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        listview.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        listview.setSpacing(2)
        self.addWidget(listview)

        selectBtn = QtWidgets.QPushButton('Select All', self)
        selectBtn.setObjectName('select_all_button')
        connect(selectBtn.clicked, self.doSelectAll)
        self.addWidget(selectBtn)

        clearBtn = QtWidgets.QPushButton('Deselect All', self)
        clearBtn.setObjectName('deselect_all_button')
        connect(clearBtn.clicked, self.doDeselectAll)
        self.addWidget(clearBtn)

        # WIDGET LAYOUT
        mainLayout = QtWidgets.QVBoxLayout(self)
        mainLayout.setAlignment(QtCore.Qt.AlignTop)
        mainLayout.addWidget(self.widget('search_field'))
        mainLayout.addWidget(self.widget('annotations_list'))
        buttonLayout = QtWidgets.QHBoxLayout()
        buttonLayout.setAlignment(QtCore.Qt.AlignLeft)
        buttonLayout.addWidget(self.widget('select_all_button'))
        buttonLayout.addWidget(self.widget('deselect_all_button'))
        mainLayout.addLayout(buttonLayout)

        self.updateModel()

    #############################################
    #   PROPERTIES
    #################################

    @property
    def project(self) -> Project:
        """
        Returns the active project.
        """
        return self._project

    #############################################
    #   SLOTS
    #################################

    @QtCore.pyqtSlot()
    def doDeselectAll(self) -> None:
        """
        Executed when the deselect all button is clicked.
        """
        for i in range(0, self.model.rowCount()):
            self.model.item(i, 0).setCheckState(QtCore.Qt.Unchecked)

    @QtCore.pyqtSlot(str)
    def doFilterItem(self, key: str = '') -> None:
        """
        Executed when the search box is filled with data.
        """
        self.proxy.setFilterRegExp(QtCore.QRegExp(key.replace(' ', '.*')))
        self.proxy.setFilterCaseSensitivity(QtCore.Qt.CaseInsensitive)
        self.proxy.sort(QtCore.Qt.AscendingOrder)

    @QtCore.pyqtSlot()
    def onReturnPressed(self) -> None:
        """
        Executed when the enter key is pressed in the search field.
        """
        self.doFilterItem()

    @QtCore.pyqtSlot()
    def doSelectAll(self) -> None:
        """
        Executed when the select all button is clicked.
        """
        for i in range(0, self.model.rowCount()):
            self.model.item(i, 0).setCheckState(QtCore.Qt.Checked)

    #############################################
    #   INTERFACE
    #################################

    def selectedAnnotations(self) -> Set[Annotation]:
        """
        Returns the set of selected annotations.
        """
        annotations = set()
        for i in range(0, self.model.rowCount()):
            item = self.model.item(i, 0)
            if item.checkState() == QtCore.Qt.Checked:
                annotations.add(item.data(self.AnnotationRole))
        return annotations

    def updateModel(self) -> None:
        """
        Update the list view model by reloading annotations from the project.
        """
        self.model.clear()
        self.model.setHorizontalHeaderLabels(['Annotation Properties'])
        for annotation in self.project.getAnnotationPropertyIRIs():
            item = QtGui.QStandardItem(str(annotation))
            item.setData(annotation, self.AnnotationRole)
            item.setCheckable(True)
            item.setCheckState(QtCore.Qt.Checked)
            self.model.appendRow(item)
        self.proxy.sort(0, QtCore.Qt.AscendingOrder)


class EntityTypesSelectionDialog(HasWidgetSystem, QtWidgets.QDialog):
    """
    Extends `QtWidgets.QDialog` providing a form for selecting a subset
    of the project annotation properties.
    """

    def __init__(
        self,
        project: 'Project',
        parent: QtWidgets.QWidget = None,
        **kwargs: Any
    ) -> None:
        """
        Initialize the dialog.
        """
        super().__init__(parent=parent, **kwargs)
        self._project = project
        self.diagrams = kwargs.get('diagrams', None)
        self.items = kwargs.get('items', {
            Item.ConceptNode,
            Item.RoleNode,
            Item.AttributeNode,
            Item.IndividualNode,
        })

        # ENTITY TYPES GROUP BOX
        groupbox = QtWidgets.QGroupBox('Entity Types')
        groupbox.setObjectName('entities_group')
        layout = QtWidgets.QHBoxLayout(groupbox)
        for item in self.items:
            checkbox = CheckBox(AbstractMetadataExporter.Types.get(item), self)
            checkbox.setObjectName(item.shortName)
            checkbox.setChecked(True)
            layout.addWidget(checkbox)
            self.addWidget(checkbox)
        self.addWidget(groupbox)

        # CONFIRMATION BOX
        confirmation = QtWidgets.QDialogButtonBox(QtCore.Qt.Horizontal, self)
        confirmation.setObjectName('confirmation_box')
        confirmation.addButton(QtWidgets.QDialogButtonBox.Cancel)
        confirmation.addButton(QtWidgets.QDialogButtonBox.Ok)
        connect(confirmation.accepted, self.accept)
        connect(confirmation.rejected, self.reject)
        self.addWidget(confirmation)

        # DIALOG LAYOUT
        mainLayout = QtWidgets.QVBoxLayout(self)
        mainLayout.setAlignment(QtCore.Qt.AlignTop)
        mainLayout.addWidget(self.widget('entities_group'))
        mainLayout.addWidget(self.widget('confirmation_box'), 0, QtCore.Qt.AlignRight)

        self.setMinimumSize(640, 100)
        self.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
        self.setWindowTitle("Select Entity Types")

    #############################################
    #   PROPERTIES
    #################################

    @property
    def project(self) -> 'Project':
        """
        Returns the active project.
        """
        return self._project

    #############################################
    #   INTERFACE
    #################################

    def selectedItems(self) -> Set[Item]:
        """
        Return the set of selected item types.
        """
        items = set()
        for item in self.items:
            checkbox = cast(QtWidgets.QCheckBox, self.widget(item.shortName))
            if checkbox.isChecked():
                items.add(item)
        return items


class AnnotationsOverridingDialog(HasWidgetSystem, QtWidgets.QDialog):
    """
    Extends QtWidgets.QDialog providing the form used to select the diagrams for a specific task like export/import
    """

    def __init__(self, session, project=None, **kwargs):
        """
        Initialize the form dialog.
        :type session: Session
        :type project: Project
        """
        super().__init__(parent=session, **kwargs)
        self._project = project
        addBtn = QtWidgets.QRadioButton('Add annotations to the existing ones', self, objectName='add_annotations', checked=True)
        overrideBtn = QtWidgets.QRadioButton('Override existing annotations', self, objectName='override_annotations')
        self.addWidget(addBtn)
        self.addWidget(overrideBtn)

        diagramLayout = QtWidgets.QGridLayout(self)
        diagramLayout.setContentsMargins(8, 8, 8, 8)

        diagramLayout.addWidget(self.widget('add_annotations'))
        diagramLayout.addWidget(self.widget('override_annotations'))

        diagramGroup = QtWidgets.QGroupBox('Overriding Options', self)
        diagramGroup.setLayout(diagramLayout)

        diagramGroupLayout = QtWidgets.QHBoxLayout(self)
        diagramGroupLayout.addWidget(diagramGroup)

        diagramWidget = QtWidgets.QWidget(self)
        diagramWidget.setLayout(diagramGroupLayout)

        confirmation = QtWidgets.QDialogButtonBox(QtCore.Qt.Horizontal, self)
        confirmation.addButton(QtWidgets.QDialogButtonBox.Ok)
        confirmation.addButton(QtWidgets.QDialogButtonBox.Cancel)
        confirmation.setObjectName('confirmation')
        self.addWidget(confirmation)

        buttonLayout = QtWidgets.QHBoxLayout(self)
        buttonLayout.setAlignment(QtCore.Qt.AlignRight)
        buttonLayout.addWidget(self.widget('confirmation'), 0, QtCore.Qt.AlignRight)

        buttonWidget = QtWidgets.QWidget(self)
        buttonWidget.setLayout(buttonLayout)

        mainLayout = QtWidgets.QVBoxLayout()
        mainLayout.setContentsMargins(10, 10, 10, 10)
        mainLayout.addWidget(diagramWidget)
        mainLayout.addWidget(buttonWidget)

        self.setLayout(mainLayout)
        self.setWindowIcon(QtGui.QIcon(':/icons/128/ic_eddy'))
        self.setWindowTitle('Overriding Options')

        connect(confirmation.accepted, self.accept)
        connect(confirmation.rejected, self.reject)

    #############################################
    #   PROPERTIES
    #################################

    @property
    def session(self):
        """
        Returns the active session (alias for self.parent()).
        :rtype: Session
        """
        return self.parent()

    @property
    def project(self):
        """
        Returns the active project.
        :rtype: Project
        """
        return self._project or self.session.project

    #############################################
    #   SLOTS
    #################################

    @QtCore.pyqtSlot()
    def checkedOption(self):

        if self.widget('override_annotations').isChecked():
            return True
        else:
            return False
